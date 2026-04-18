"""
utils.py – Export, import en KPI-hulpfuncties
Ondersteunt: Excel, CSV, iCal, PDF (ReportLab), budget-berekeningen
"""
from __future__ import annotations
import calendar
import io
import json
import uuid
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple

# ─── Optionele imports ────────────────────────────────────────────────────────

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    from icalendar import Calendar, Event as ICalEvent
    HAS_ICAL = True
except ImportError:
    HAS_ICAL = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    )
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

from data_model import Employee, Shift, Schedule, AppState
from constraints import dutch_holidays, check_schedule


# ─── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def export_schedule_excel(schedule: Schedule, employees: List[Employee]) -> bytes:
    """Exporteer maandrooster als Excel-bestand."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl niet geïnstalleerd.")

    wb = openpyxl.Workbook()

    # ── Blad 1: Maandoverzicht ─────────────────────────────────────────────
    ws_month = wb.active
    ws_month.title = "Maandrooster"
    year, month = schedule.year, schedule.month
    days_in_month = calendar.monthrange(year, month)[1]
    nl_days = ["Ma", "Di", "Wo", "Do", "Vr", "Za", "Zo"]
    holidays = dutch_holidays(year)

    # Header rij
    ws_month.cell(1, 1, "Medewerker").font = Font(bold=True)
    ws_month.cell(1, 2, "Contractu.").font = Font(bold=True)
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        col = day + 2
        cell = ws_month.cell(1, col, f"{nl_days[d.weekday()]} {day}")
        cell.font = Font(bold=True)
        # Kleur weekend / feestdag
        if d.weekday() == 6:
            cell.fill = PatternFill("solid", fgColor="FFD700")
        elif d.weekday() == 5:
            cell.fill = PatternFill("solid", fgColor="FFF3CD")
        if str(d) in holidays:
            cell.fill = PatternFill("solid", fgColor="FFCCCC")

    # Medewerker rijen
    emp_map = {e.id: e for e in employees}
    emp_shifts: Dict[str, Dict[str, List[Shift]]] = {}
    for sh in schedule.shifts:
        emp_shifts.setdefault(sh.employee_id, {}).setdefault(sh.date, []).append(sh)

    for row, emp in enumerate(employees, start=2):
        ws_month.cell(row, 1, emp.name)
        ws_month.cell(row, 2, f"{emp.contract_hours}u")
        for day in range(1, days_in_month + 1):
            d = date(year, month, day)
            date_str = str(d)
            shifts_today = emp_shifts.get(emp.id, {}).get(date_str, [])
            if shifts_today:
                sh = shifts_today[0]
                cell = ws_month.cell(row, day + 2, f"{sh.start_time}-{sh.end_time}")
                cell.fill = PatternFill("solid", fgColor="C8E6C9")
            else:
                ws_month.cell(row, day + 2, "")

    ws_month.column_dimensions["A"].width = 20
    ws_month.column_dimensions["B"].width = 10
    for col in range(3, days_in_month + 3):
        ws_month.column_dimensions[get_column_letter(col)].width = 11

    # ── Blad 2: Diensten detail ────────────────────────────────────────────
    ws_detail = wb.create_sheet("Diensten Detail")
    headers = ["Datum", "Dag", "Medewerker", "Functie", "Begin", "Einde",
               "Uren", "Pauze (min)", "Toeslag", "Loonkosten €", "Notities"]
    for col, h in enumerate(headers, 1):
        cell = ws_detail.cell(1, col, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2196F3")
        cell.font = Font(bold=True, color="FFFFFF")

    for row, sh in enumerate(sorted(schedule.shifts, key=lambda s: s.date), start=2):
        d = date.fromisoformat(sh.date)
        supplement = []
        if sh.has_sunday_supplement:   supplement.append("Za+")
        if sh.has_holiday_supplement:  supplement.append("FD+")
        if sh.is_night_shift:          supplement.append("Nacht+")
        ws_detail.cell(row, 1, sh.date)
        ws_detail.cell(row, 2, nl_days[d.weekday()])
        ws_detail.cell(row, 3, sh.employee_name)
        ws_detail.cell(row, 4, sh.role)
        ws_detail.cell(row, 5, sh.start_time)
        ws_detail.cell(row, 6, sh.end_time)
        ws_detail.cell(row, 7, round(sh.duration_hours, 2))
        ws_detail.cell(row, 8, sh.break_minutes)
        ws_detail.cell(row, 9, ", ".join(supplement) if supplement else "–")
        ws_detail.cell(row, 10, round(sh.labor_cost, 2))
        ws_detail.cell(row, 11, sh.notes)

    # ── Blad 3: KPI-samenvatting ──────────────────────────────────────────
    ws_kpi = wb.create_sheet("KPI Samenvatting")
    kpis = compute_kpis(schedule, employees)
    ws_kpi.cell(1, 1, "KPI").font = Font(bold=True)
    ws_kpi.cell(1, 2, "Waarde").font = Font(bold=True)
    for row, (k, v) in enumerate(kpis.items(), start=2):
        ws_kpi.cell(row, 1, k)
        ws_kpi.cell(row, 2, str(v))
    ws_kpi.column_dimensions["A"].width = 35
    ws_kpi.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_employees_csv(employees: List[Employee]) -> str:
    """Exporteer medewerkerslijst als CSV."""
    lines = ["id,naam,contracturen,salaristype,uurloon,maandsalaris,functie,senioriteit,startdatum,email,telefoon"]
    for e in employees:
        lines.append(
            f"{e.id},{e.name},{e.contract_hours},{e.salary_type},{e.hourly_rate},"
            f"{e.monthly_salary},{e.role},{e.seniority_years},{e.start_date},{e.email},{e.phone}"
        )
    return "\n".join(lines)


def import_employees_csv(content: str) -> Tuple[List[Employee], List[str]]:
    """Importeer medewerkers uit CSV. Geeft (employees, errors) terug."""
    from data_model import Employee, EmployeeAvailability
    employees, errors = [], []
    lines = content.strip().split("\n")
    if not lines:
        return employees, ["Leeg bestand"]
    headers = [h.strip().lower() for h in lines[0].split(",")]
    for i, line in enumerate(lines[1:], start=2):
        if not line.strip():
            continue
        parts = line.split(",")
        if len(parts) < 3:
            errors.append(f"Rij {i}: te weinig kolommen")
            continue
        try:
            emp = Employee()
            # Probeer velden te mappen
            field_map = {
                "naam": "name", "name": "name",
                "contracturen": "contract_hours", "contract_hours": "contract_hours",
                "uurloon": "hourly_rate", "hourly_rate": "hourly_rate",
                "maandsalaris": "monthly_salary",
                "salaristype": "salary_type",
                "functie": "role", "role": "role",
                "startdatum": "start_date", "start_date": "start_date",
                "email": "email", "telefoon": "phone", "phone": "phone",
            }
            for col_idx, header in enumerate(headers):
                if col_idx >= len(parts):
                    break
                attr = field_map.get(header)
                val = parts[col_idx].strip()
                if attr and hasattr(emp, attr):
                    try:
                        existing = getattr(emp, attr)
                        if isinstance(existing, float):
                            setattr(emp, attr, float(val) if val else existing)
                        elif isinstance(existing, int):
                            setattr(emp, attr, int(val) if val else existing)
                        else:
                            setattr(emp, attr, val)
                    except (ValueError, TypeError):
                        pass
            employees.append(emp)
        except Exception as ex:
            errors.append(f"Rij {i}: {ex}")
    return employees, errors


# ─── ICAL EXPORT ──────────────────────────────────────────────────────────────

def export_ical_employee(shifts: List[Shift], emp_name: str) -> bytes:
    """Genereer iCal-bestand voor één medewerker."""
    if not HAS_ICAL:
        raise ImportError("icalendar niet geïnstalleerd.")
    cal = Calendar()
    cal.add("prodid", f"-//McDonald's Rooster//{emp_name}//NL")
    cal.add("version", "2.0")
    cal.add("calscale", "GREGORIAN")
    cal.add("x-wr-calname", f"Rooster {emp_name}")

    for sh in shifts:
        event = ICalEvent()
        d     = date.fromisoformat(sh.date)
        sh_h, sh_m = map(int, sh.start_time.split(":"))
        eh, em     = map(int, sh.end_time.split(":"))
        start_dt   = datetime(d.year, d.month, d.day, sh_h, sh_m)
        end_m_total = eh * 60 + em
        start_m     = sh_h * 60 + sh_m
        if end_m_total <= start_m:
            end_m_total += 1440
        extra = end_m_total // 1440
        final = end_m_total % 1440
        end_d = d + timedelta(days=extra)
        end_dt = datetime(end_d.year, end_d.month, end_d.day, final // 60, final % 60)

        event.add("uid", f"{sh.id}@mcdonalds-rooster")
        event.add("summary", f"Dienst {sh.role} – {sh.start_time} t/m {sh.end_time}")
        event.add("dtstart", start_dt)
        event.add("dtend", end_dt)
        event.add("description",
                  f"Functie: {sh.role}\nPauze: {sh.break_minutes} min\n{sh.notes}")
        cal.add_component(event)

    return cal.to_ical()


def export_ical_full(schedule: Schedule) -> bytes:
    """Genereer iCal voor het volledige team."""
    if not HAS_ICAL:
        raise ImportError("icalendar niet geïnstalleerd.")
    cal = Calendar()
    cal.add("prodid", f"-//McDonald's Rooster//Team//NL")
    cal.add("version", "2.0")
    cal.add("calscale", "GREGORIAN")
    cal.add("x-wr-calname",
            f"Teamrooster {schedule.restaurant.name} {schedule.month}/{schedule.year}")

    for sh in schedule.shifts:
        event = ICalEvent()
        d     = date.fromisoformat(sh.date)
        sh_h, sh_m = map(int, sh.start_time.split(":"))
        event.add("uid", f"{sh.id}@mcdonalds-rooster-team")
        event.add("summary", f"{sh.employee_name}: {sh.start_time}-{sh.end_time}")
        event.add("dtstart", datetime(d.year, d.month, d.day, sh_h, sh_m))
        cal.add_component(event)

    return cal.to_ical()


# ─── PDF EXPORT ───────────────────────────────────────────────────────────────

def export_schedule_pdf(schedule: Schedule, employees: List[Employee]) -> bytes:
    """Genereer PDF-rooster (A4 landscape) via ReportLab."""
    if not HAS_REPORTLAB:
        raise ImportError("reportlab niet geïnstalleerd.")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=15*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    story  = []
    year, month = schedule.year, schedule.month
    days_in_month = calendar.monthrange(year, month)[1]
    nl_months = ["", "Januari", "Februari", "Maart", "April", "Mei", "Juni",
                 "Juli", "Augustus", "September", "Oktober", "November", "December"]
    nl_days   = ["Ma", "Di", "Wo", "Do", "Vr", "Za", "Zo"]
    holidays  = dutch_holidays(year)

    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
    story.append(Paragraph(
        f"🍟 {schedule.restaurant.name} – Rooster {nl_months[month]} {year}", title_style))
    story.append(Paragraph(
        f"Gegenereerd op: {datetime.now().strftime('%d-%m-%Y %H:%M')} | "
        f"Budget: €{schedule.labor_budget:,.0f}", styles["Normal"]))
    story.append(Spacer(1, 6*mm))

    # Bouw tabeldata
    header_row = ["Medewerker"] + [
        f"{nl_days[date(year, month, d).weekday()]}\n{d}"
        for d in range(1, days_in_month + 1)
    ] + ["Totaal"]

    emp_shifts: Dict[str, Dict[str, List[Shift]]] = {}
    for sh in schedule.shifts:
        emp_shifts.setdefault(sh.employee_id, {}).setdefault(sh.date, []).append(sh)

    table_data = [header_row]
    for emp in employees:
        row = [emp.name]
        total_h = 0.0
        for day in range(1, days_in_month + 1):
            d        = date(year, month, day)
            date_str = str(d)
            s_list   = emp_shifts.get(emp.id, {}).get(date_str, [])
            if s_list:
                sh = s_list[0]
                row.append(f"{sh.start_time}\n{sh.end_time}")
                total_h += sh.duration_hours
            else:
                row.append("")
        row.append(f"{total_h:.1f}u")
        table_data.append(row)

    col_w = [35*mm] + [7.5*mm] * days_in_month + [12*mm]
    table = Table(table_data, colWidths=col_w, repeatRows=1)

    # Stijlen
    style_cmds = [
        ("BACKGROUND",  (0, 0), (-1, 0), colors.HexColor("#DD0000")),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 6),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.grey),
        ("FONTNAME",    (0, 1), (0, -1), "Helvetica-Bold"),
    ]
    # Kleur weekendkolommen
    for day in range(1, days_in_month + 1):
        d = date(year, month, day)
        col = day  # kolom index (0 = naam)
        if d.weekday() == 6:
            style_cmds.append(("BACKGROUND", (col, 0), (col, -1), colors.HexColor("#FFD700")))
        elif str(d) in holidays:
            style_cmds.append(("BACKGROUND", (col, 0), (col, -1), colors.HexColor("#FFCCCC")))

    table.setStyle(TableStyle(style_cmds))
    story.append(table)

    # KPI samenvatting
    story.append(PageBreak())
    story.append(Paragraph("KPI Samenvatting", title_style))
    kpis = compute_kpis(schedule, employees)
    kpi_data = [["KPI", "Waarde"]] + [[k, str(v)] for k, v in kpis.items()]
    kpi_table = Table(kpi_data, colWidths=[120*mm, 60*mm])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DD0000")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(kpi_table)

    doc.build(story)
    return buf.getvalue()


# ─── KPI BEREKENINGEN ─────────────────────────────────────────────────────────

def compute_kpis(schedule: Schedule, employees: List[Employee]) -> Dict:
    """Bereken KPI's voor het dashboard."""
    year, month = schedule.year, schedule.month
    days_in_month = calendar.monthrange(year, month)[1]
    weeks = days_in_month / 7
    holidays = dutch_holidays(year)

    total_cost     = sum(sh.labor_cost for sh in schedule.shifts)
    total_hours    = sum(sh.duration_hours for sh in schedule.shifts)
    budget         = schedule.labor_budget
    budget_var     = total_cost - budget
    budget_pct     = (total_cost / budget * 100) if budget > 0 else 0

    # Uren per medewerker vs contract
    emp_hours: Dict[str, float] = {}
    for sh in schedule.shifts:
        emp_hours[sh.employee_id] = emp_hours.get(sh.employee_id, 0) + sh.duration_hours

    contract_deviations = []
    overtime_hours      = 0.0
    for emp in employees:
        target = emp.contract_hours * weeks
        actual = emp_hours.get(emp.id, 0.0)
        dev    = actual - target
        contract_deviations.append(abs(dev))
        if dev > 0:
            overtime_hours += dev

    avg_deviation   = sum(contract_deviations) / len(contract_deviations) if contract_deviations else 0
    overtime_pct    = (overtime_hours / total_hours * 100) if total_hours > 0 else 0

    # Ziekteverzuim
    total_sick_days = sum(len(e.sick_log) for e in employees)
    working_days    = sum(1 for d in range(1, days_in_month + 1)
                         if date(year, month, d).weekday() < 5)
    sick_pct        = (total_sick_days / (len(employees) * working_days) * 100) if employees and working_days else 0

    # Shifts per dag gemiddeld
    shifts_per_day  = len(schedule.shifts) / days_in_month if days_in_month > 0 else 0

    # Feestdagen
    n_holidays = sum(1 for d in range(1, days_in_month + 1)
                     if str(date(year, month, d)) in holidays)

    # Constraint-schendingen
    violations = check_schedule(schedule, employees)
    n_errors   = sum(1 for v in violations if v.severity == "error")
    n_warnings = sum(1 for v in violations if v.severity == "warning")

    return {
        "Totale loonkosten":           f"€ {total_cost:,.2f}",
        "Budget":                      f"€ {budget:,.2f}",
        "Budget-afwijking":            f"€ {budget_var:+,.2f} ({budget_pct:.1f}%)",
        "Totale diensturen":           f"{total_hours:.1f} uur",
        "Gem. uren/dag":               f"{shifts_per_day:.1f} diensten",
        "Overtijduren":                f"{overtime_hours:.1f} uur ({overtime_pct:.1f}%)",
        "Gem. contractafwijking/emp":  f"{avg_deviation:.1f} uur",
        "Ziekteverzuim (maand)":       f"{sick_pct:.1f}%",
        "Feestdagen in maand":         str(n_holidays),
        "CAO-overtredingen":           f"{n_errors} fouten, {n_warnings} waarschuwingen",
        "Actieve medewerkers":         str(len(employees)),
    }


# ─── GANTT DATA ───────────────────────────────────────────────────────────────

def build_gantt_data(schedule: Schedule) -> List[Dict]:
    """Bouw data voor Plotly Gantt-chart."""
    import pandas as pd  # noqa
    rows = []
    for sh in schedule.shifts:
        d  = date.fromisoformat(sh.date)
        sh_h, sh_m = map(int, sh.start_time.split(":"))
        eh, em     = map(int, sh.end_time.split(":"))
        start_dt   = datetime(d.year, d.month, d.day, sh_h, sh_m)
        end_m_total = eh * 60 + em
        start_m     = sh_h * 60 + sh_m
        if end_m_total <= start_m:
            end_m_total += 1440
        extra = end_m_total // 1440
        final = end_m_total % 1440
        end_d = d + timedelta(days=extra)
        end_dt = datetime(end_d.year, end_d.month, end_d.day, final // 60, final % 60)

        color = "#4CAF50"
        if sh.is_night_shift:      color = "#9C27B0"
        if sh.has_holiday_supplement: color = "#F44336"
        if sh.has_sunday_supplement:  color = "#FF9800"

        rows.append({
            "Task":   sh.employee_name,
            "Start":  start_dt.isoformat(),
            "Finish": end_dt.isoformat(),
            "Resource": sh.role,
            "Color":  color,
            "Shift ID": sh.id,
            "Kosten": f"€{sh.labor_cost:.2f}",
            "Label":  f"{sh.start_time}-{sh.end_time}",
        })
    return rows


# ─── BEWAAR / LAAD STATE ─────────────────────────────────────────────────────

def state_to_json(state: "AppState") -> str:
    return json.dumps(state.to_dict(), ensure_ascii=False, indent=2)


def state_from_json(json_str: str) -> "AppState":
    from data_model import AppState
    data = json.loads(json_str)
    return AppState.from_dict(data)
